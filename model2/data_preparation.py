import torch
from transformers import T5ForConditionalGeneration, T5Tokenizer
import openpyxl
import json
import re
import sys
import os

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

model_path = "../model/agriculture_text_transform_model"
tokenizer = T5Tokenizer.from_pretrained(model_path)
model = T5ForConditionalGeneration.from_pretrained(model_path).cuda()
model.eval()


def transform_text(text):
    inputs = tokenizer(
        text,
        return_tensors='pt',
        truncation=True,
        max_length=128
    ).to(model.device)

    with torch.no_grad():
        outputs = model.generate(
            **inputs,
            max_length=128,
            num_beams=5,
            repetition_penalty=2.5,
            early_stopping=True
        )

    return tokenizer.decode(outputs[0], skip_special_tokens=True)


operations = [
    "Гербицидная обработка", "1 Гербицидная обработка", "2 Гербицидная обработка", "Инсектицидная обработка", "3 Гербицидная обработка",
    "4 Гербицидная обработка", "2-е Выравнивание зяби", "Выравнивание зяби", "Внесение минеральных удобрений", "Боронование довсходовое",
    "2-я междурядная культивация", "1-я междурядная культивация", "Предпосевная культивация", "Прикатывание посевов", "Выравнивание зяби",
    "Гербицидная обработка", "Сев", "Уборка", "Культивация", "Боронование", "Сплошная культивация", "2-е Дискование",
    "Инсектицидная обработка", "Пахота", "Дискование", "Подкормка", "Функицидная обработка", "Чизлевание", "Затравка мышевидных грызунов"
]

crops = [
    "Пшеница озимая", "Пшеница озимая товарная", "Пшеница озимая семенная", "Кукуруза", "Гречиха", "Рапс",
    "Ячмень озимый", "Подсолнечник", "Свекла сахарная", "Соя", "Многолетние травы", "Вика+Тритикале",
    "Горох на зерно", "Горох товарный", "Гуар", "Конопля", "Кориандр", "Кукуруза кормовая",
    "Кукуруза семенная", "Кукуруза товарная", "Люцерна", "Многолетние злаковые травы", "Многолетние травы прошлых лет",
    "Многолетние травы текущего года", "Овес", "Подсолнечник кондитерский", "Подсолнечник семенной",
    "Подсолнечник товарный", "Просо", "Пшеница озимая на зеленый корм", "Пшеница озимая семенная",
    "Пшеница озимая товарная", "Рапс озимый", "Рапс яровой", "Свекла сахарная", "Сорго", "Сорго кормовой",
    "Сорго-суданковый гибрид", "Соя семенная", "Соя товарная", "Чистый пар", "Чумиза", "Ячмень озимый", "Ячмень озимый семенной"
]

departments = [
    "АОР", "Кавказ", "Север", "Центр", "Юг", "ТСК", "АО Кропоткинское", "Восход", "Колхоз Прогресс", "Мир", "СП Коломейцево"
]

def tokenize_and_label(text):
    tokens = text.split()

    ner_tags = ["O"] * len(tokens)

    date_pattern = re.compile(r"^\d{2}\.\d{2}$")
    yield_total_pattern = re.compile(r"^\d+(?:/\d+)?$")

    for i, token in enumerate(tokens):
        for op in operations:
            if " ".join(tokens[i:i + len(op.split())]) == op:
                for j in range(i, i + len(op.split())):
                    ner_tags[j] = "B-OPERATION" if j == i else "I-OPERATION"

        for crop in crops:
            if " ".join(tokens[i:i + len(crop.split())]) == crop:
                for j in range(i, i + len(crop.split())):
                    ner_tags[j] = "B-CROP" if j == i else "I-CROP"

        if date_pattern.match(token):
            ner_tags[i] = "B-DATE"

        if token.lower() == "вал" and i + 1 < len(tokens):
            if yield_total_pattern.match(tokens[i + 1]):
                ner_tags[i] = "B-YIELD_TOTAL"
                ner_tags[i + 1] = "I-YIELD_TOTAL"

        departments_lower = [dept.lower() for dept in departments]
        base_token = token.split('-')[0].split(':')[0].lower()

        if base_token.lower() in departments_lower:
            ner_tags[i] = "B-DEPARTMENT"

    for i in range(len(tokens) - 2):
        if tokens[i].lower() == "по" and tokens[i + 1].lower() == "пу":
            ner_tags[i] = "B-SUBUNIT"
            ner_tags[i + 1] = "B-SUBUNIT"
            if re.match(r"\d+/\d+", tokens[i + 2]):
                ner_tags[i + 2] = "I-SUBUNIT"

    for i in range(len(tokens) - 1):
        if tokens[i].lower().startswith("отд"):
            ner_tags[i] = "B-DEPARTMENT"
            if re.match(r"^\d+$", tokens[i + 1]):
                ner_tags[i + 1] = "I-DEPARTMENT"

    result = {
        "tokens": tokens,
        "ner_tags": ner_tags
    }

    return result

wb = openpyxl.load_workbook("train.xlsx")
ws = wb.active

text_column = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row + 1)]
processed_data = []

for text in text_column:
    # agriculture_text_transform_model
    generated_text = transform_text(text)
    # print(generated_text)

    # Разметка NER
    processed_data.append(tokenize_and_label(generated_text))

processed_data = [tokenize_and_label(text) for text in text_column]

with open("processed_data.json", "w", encoding="utf-8") as f:
    json.dump(processed_data, f, ensure_ascii=False, indent=4)

print("0")