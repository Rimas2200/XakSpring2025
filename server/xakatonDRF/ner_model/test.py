import torch
from transformers import BertTokenizerFast, BertForTokenClassification
from openpyxl import load_workbook, Workbook
import os
import re
import sys
from transformers import T5ForConditionalGeneration, T5Tokenizer

# Указываем путь к локальной модели
# model_path = os.path.getcwd() + "/T5_model/agriculture_text_transform_model"
model_path = os.path.join(os.getcwd(), "T5_model", "agriculture_text_transform_model")
# Проверяем, существует ли путь
print(model_path)
if not os.path.exists(model_path):
    raise FileNotFoundError(f"Модель не найдена по пути: {model_path}")

# Загружаем токенайзер и модель с локального пути
T5tokenizer = T5Tokenizer.from_pretrained(model_path, local_files_only=True)
T5model = T5ForConditionalGeneration.from_pretrained(model_path, local_files_only=True).to(torch.device('cpu') )
T5model.eval()


def transform_text(text):
    inputs = T5tokenizer(
        text,
        return_tensors='pt',
        truncation=True,
        max_length=128
    ).to(T5model.device)

    with torch.no_grad():
        outputs = T5model.generate(
            **inputs,
            max_length=128,
            num_beams=5,
            repetition_penalty=2.5,
            early_stopping=True
        )

    return T5tokenizer.decode(outputs[0], skip_special_tokens=True)

def preprocess_with_t5(text):
    transformed_text = transform_text(text)
    return transformed_text

model_path = os.path.join(os.getcwd(), "ner_model", "ner-model")
tokenizer_ner = BertTokenizerFast.from_pretrained(model_path)
model_ner = BertForTokenClassification.from_pretrained(model_path)

"""
O - символы, которые нахуй не нужны # не нужная информация

B-OPERATION - начало наименование операции
I-OPERATION - часть операции

B-CROP - начало наименования культуры
I-CROP - часть культуры

B-DATE - дата

B-SUBUNIT - начинает наименование пу
I-SUBUNIT - наименование пу

B-DEPARTMENT - начинает наименование отделения
I-DEPARTMENT - наименование отделения

B-YIELD_TOTAL - вал и все с ним связанное

B-HECTARE - га
I-HECTARE - га
"""
label_list = [
    'B-CROP', 
    'B-DATE', 
    'B-DEPARTMENT', 
    'B-HECTARE', 
    'B-OPERATION', 
    'B-SUBUNIT', 
    'B-YIELD_TOTAL', 
    
    'I-CROP', 
    'I-DEPARTMENT', 
    'I-HECTARE', 
    'I-OPERATION', 
    'I-SUBUNIT', 
    'I-YIELD_TOTAL', 
    'O'
]



department_mapping = {
    "1": "АОР",
    "3": "АОР",
    "4": "АОР",
    "5": "АОР",
    "6": "АОР",
    "7": "АОР",
    "9": "АОР",
    "10": "АОР",
    "11": "АОР",
    "12": "АОР",
    "16": "АОР",
    "17": "АОР",
    "18": "АОР",
    "19": "АОР",
    "20": "АОР",
    "Кавказ": "АОР",
    "Север": "АОР",
    "Центр": "АОР",
    "Юг": "АОР",
}

def remove_duplicate_words(text):
    seen = set()
    result = []
    for word in text.split():
        if word not in seen:
            seen.add(word)
            result.append(word)
    return " ".join(result)


def predict_entities(text):
    words = text.split()

    inputs = tokenizer_ner(
        words,
        is_split_into_words=True,
        return_tensors="pt",
        truncation=True,
        padding=True,
        max_length=64
    )

    with torch.no_grad():
        outputs = model_ner(**inputs)

    predictions = torch.argmax(outputs.logits, dim=2)[0].tolist()

    word_ids = inputs.word_ids()
    entities = []
    current_entity = None
    current_entity_words = []
    current_entity_start = None
    last_word_idx = None

    for word_idx, label_id in zip(word_ids, predictions):
        if word_idx is None:
            continue

        last_word_idx = word_idx

        label = label_list[label_id]
        word = words[word_idx]

        if label.startswith("B-"):
            if current_entity:
                entities.append({
                    "entity": current_entity,
                    "text": " ".join(current_entity_words),
                    "start": current_entity_start,
                    "end": word_idx
                })
            current_entity = label[2:]
            current_entity_words = [word]
            current_entity_start = word_idx

        elif label.startswith("I-") and current_entity and label[2:] == current_entity:
            current_entity_words.append(word)

        else:
            if current_entity:
                entities.append({
                    "entity": current_entity,
                    "text": " ".join(current_entity_words),
                    "start": current_entity_start,
                    "end": word_idx
                })
                current_entity = None
                current_entity_words = []
                current_entity_start = None

    if current_entity and last_word_idx is not None:
        entities.append({
            "entity": current_entity,
            "text": " ".join(current_entity_words),
            "start": current_entity_start,
            "end": last_word_idx + 1
        })

    merged_entities = []
    for entity in entities:
        if not merged_entities:
            merged_entities.append(entity)
        else:
            last_entity = merged_entities[-1]
            if (last_entity['entity'] == entity['entity'] and
                    last_entity['end'] == entity['start']):
                last_entity['text'] += " " + entity['text']
                last_entity['end'] = entity['end']
            else:
                merged_entities.append(entity)

    for entity in merged_entities:
        entity['text'] = remove_duplicate_words(entity['text'])

    return merged_entities


def process_subunit_and_hectare(entities):
    hectare_values = []

    for entity in entities:
        key = entity['entity'].upper()
        if key == "HECTARE":
            match = re.search(r'(\d+)', entity['text'])
            if match:
                hectare_values.append(int(match.group(1)))

    for entity in entities:
        key = entity['entity'].upper()
        if key == "SUBUNIT":
            subunit_value = entity['text']
            match = re.search(r'по\s*пу\s*(\d+)/(\d+)', subunit_value, re.IGNORECASE)
            if match:
                part1, part2 = int(match.group(1)), int(match.group(2))
                entities = [e for e in entities if e['entity'].upper() != "SUBUNIT"]
                entities.append({"entity": "HECTARE", "text": str(part1)})
                entities.append({"entity": "SUBUNIT", "text": str(part2)})
                break

    if len(hectare_values) >= 1:
        smaller = min(hectare_values)
        larger = max(hectare_values)
        entities.append({"entity": "HECTARE", "text": str(smaller)})
        entities.append({"entity": "SUBUNIT", "text": str(larger)})
        print(smaller, larger)
    elif len(hectare_values) == 1:
        print("------------------")
        entities.append({"entity": "HECTARE", "text": str(hectare_values[0])})

    return entities


def group_entities_by_operation(entities):
    """
    Группирует сущности по операциям.
    """
    grouped_data = []
    current_group = {}
    global_data = {"DATE": "", "DEPARTMENT": ""}

    for entity in entities:
        key = entity['entity'].upper()

        if key == "OPERATION":
            if current_group:
                grouped_data.append(current_group)
                current_group = {}
            current_group[key] = entity['text']
        elif key in ["DATE", "DEPARTMENT"]:
            global_data[key] = entity['text']
        elif key in ["CROP", "HECTARE", "SUBUNIT", "YIELD_TOTAL", "YIELD_TOTAL_TOTAL"]:
            current_group[key] = entity['text']

    if current_group:
        grouped_data.append(current_group)

    for group in grouped_data:
        group.update(global_data)

    return grouped_data

def process_department(entities):
    """
    Обрабатывает значение DEPARTMENT
    """
    for entity in entities:
        key = entity['entity'].upper()
        if key == "DEPARTMENT":
            department_value = entity['text']
            if department_value in department_mapping:
                entity['text'] = department_mapping[department_value]
            elif "отд" in department_value.lower():
                match = re.search(r'\d+', department_value)
                if match:
                    department_number = match.group(0)
                    if department_number in department_mapping:
                        entity['text'] = department_mapping[department_number]
            elif "юг" in department_value.lower():
                entity['text'] = "АОР"
    return entities


import re

def process_yield_total(entities):
    """
    Обрабатывает значение YIELD_TOTAL.
    Если значение содержит два числа через слеш, разделяет их на два поля:
    - YIELD_TOTAL_DAY (первое число)
    - YIELD_TOTAL_TOTAL (второе число).
    Если значение содержит одно число, оно записывается только в YIELD_TOTAL_DAY.
    """
    for entity in entities:
        key = entity['entity'].upper()
        if key == "YIELD_TOTAL":
            yield_total_value = entity['text']
            matches = re.findall(r'\d+', yield_total_value)
            if len(matches) == 2:
                entity['text'] = matches[0]
                entities.append({'entity': 'YIELD_TOTAL_TOTAL', 'text': matches[1]})
            elif len(matches) == 1:
                entity['text'] = matches[0]
    return entities


def write_to_excel(entities, file_name="Таблица (полевые работы).xlsx"):
    headers = [
        "Дата", "Подразделение", "Операция", "Культура",
        "За день, га", "С начала операции, га", "Вал за день, ц", "Вал с начала, ц"
    ]

    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)

    entity_to_column = {
        "DATE": "Дата",
        "DEPARTMENT": "Подразделение",
        "OPERATION": "Операция",
        "CROP": "Культура",
        "HECTARE": "За день, га",
        "SUBUNIT": "С начала операции, га",
        "YIELD_TOTAL": "Вал за день, ц",
        "YIELD_TOTAL_TOTAL": "Вал с начала, ц"
    }

    grouped_data = group_entities_by_operation(entities)

    for group in grouped_data:
        row_data = {header: "" for header in headers}
        for key, value in group.items():
            if key in entity_to_column:
                col_name = entity_to_column[key]
                row_data[col_name] = value
        sheet.append([row_data[header] for header in headers])
    workbook.save(file_name)

def process_file(input_file_path, output_file_name="Таблица (полевые работы).xlsx"):
    """
    Обрабатывает файл построчно и записывает результаты в Excel.
    :param input_file_path: Путь к входному текстовому файлу.
    :param output_file_name: Имя выходного Excel-файла.
    """

    with open(input_file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    for line in lines:
        line = line.strip()
        if not line:
            continue

        print(f"Обработка строки: {line}")

        transformed_text = preprocess_with_t5(line)
        print(f"Трансформированный текст: {transformed_text}")

        entities = predict_entities(transformed_text)
        print(f"Извлеченные сущности: {entities}")

        entities = process_subunit_and_hectare(entities)
        entities = process_department(entities)
        entities = process_yield_total(entities)

        write_to_excel(entities, output_file_name)


def process_file_txt(input_file_path, output_file_name="Таблица.txt"):
    """
    Обрабатывает файл построчно и записывает результаты в txt.
    :param input_file_path: Путь к входному текстовому файлу.
    :param output_file_name: Имя выходного txt-файла.
    """
    with open(input_file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    with open(output_file_name, 'w', encoding='utf-8') as output_file:
        for line in lines:
            line = line.strip()
            if not line:
                continue
            transformed_text = preprocess_with_t5(line)
            output_file.write(f"{transformed_text}\n")

    print(f"Обработка завершена. Результаты сохранены в файл: {output_file_name}")

def split_operations(input_file_path, output_file_name="Таблица_разбитая.txt"):
    """
    Разделяет строки из входного файла на отдельные операции и записывает их в выходной файл.
    :param input_file_path: Путь к входному текстовому файлу.
    :param output_file_name: Имя выходного txt-файла.
    """
    with open(input_file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    with open(output_file_name, 'w', encoding='utf-8') as output_file:
        for line in lines:
            line = line.strip()
            if not line:
                continue

            date_match = re.match(r"\d{2}\.\d{2}", line)
            if date_match:
                date = date_match.group(0)
                rest_of_line = line[len(date):].strip()
            else:
                date = ""
                rest_of_line = line

            key_words = re.findall(r"(АОР|ТСК|АО Кропоткинское|Восход|Колхоз Прогресс|Мир|СП Коломейцево)", rest_of_line)
            key_word = key_words[0] if key_words else ""

            rest_of_line = re.sub(r"(АОР|ТСК|АО Кропоткинское|Восход|Колхоз Прогресс|Мир|СП Коломейцево)", "", rest_of_line).strip()

            operations = re.split(r"(?=Пахота|Выравнивание|Сев|Культивация|Подкормка|Внесение|Уборка|Предпосевная|Чизлевание|Химпрополка|Первая|Сплошная|2-е)", rest_of_line)

            for operation in operations:
                if operation.strip():
                    output_file.write(f"{date} {key_word} {operation.strip()}\n")

    print(f"Обработка завершена. Результаты сохранены в файл: {output_file_name}")


def remove_rows_without_operation_or_crop(file_name="Таблица (полевые работы).xlsx"):
    if not os.path.exists(file_name):
        raise FileNotFoundError(f"Файл {file_name} не найден.")

    workbook = load_workbook(file_name)
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]
    operation_col_index = headers.index("Операция") + 1
    crop_col_index = headers.index("Культура") + 1

    rows_to_delete = []

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
        operation_value = row[operation_col_index - 1].value
        crop_value = row[crop_col_index - 1].value

        if not operation_value or not crop_value:
            rows_to_delete.append(row_idx)

    for row_idx in reversed(rows_to_delete):
        sheet.delete_rows(row_idx)

    workbook.save(file_name)


if __name__ == "__main__":# output_file_name = "Таблица (полевые работы).xlsx"
    process_file_txt("message.txt", "Таблица.txt")
    split_operations("Таблица.txt", "Таблица_разбитая.txt")
    process_file("Таблица_разбитая.txt", "Таблица (полевые работы).xlsx")
    remove_rows_without_operation_or_crop("Таблица (полевые работы).xlsx")
